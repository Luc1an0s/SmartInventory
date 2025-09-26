import sqlite3
import pandas as pd
import smtplib
from email.message import EmailMessage
import os
from fpdf import FPDF
from openpyxl import load_workbook
from openpyxl.styles import Font

conn = sqlite3.connect("estoque.db")
produtos_df = pd.read_sql("SELECT * FROM produtos", conn)
vendas_df = pd.read_sql("SELECT * FROM vendas", conn)
conn.close()

vendas_df['data'] = pd.to_datetime(vendas_df['data'])
vendas_por_dia = vendas_df.groupby(['loja_id', 'produto_id', 'data']).sum().reset_index()
media_diaria = vendas_por_dia.groupby(['loja_id', 'produto_id'])['quantidade'].mean().reset_index()
media_diaria.rename(columns={'quantidade': 'media_diaria'}, inplace=True)

estoque_df = produtos_df.rename(columns={'id': 'produto_id', 'quantidade': 'estoque_atual'})
estoque_df = estoque_df.merge(media_diaria, on=['loja_id', 'produto_id'], how='left')
estoque_df['media_diaria'] = estoque_df['media_diaria'].fillna(0)

faltando = estoque_df[estoque_df['estoque_atual'] < estoque_df['estoque_minimo']]
relatorio = []

lojas_com_falta = faltando['loja_id'].unique()

for loja_destino in lojas_com_falta:
    produtos_faltando = faltando[faltando['loja_id'] == loja_destino]

    for _, falta in produtos_faltando.iterrows():
        produto_id = falta['produto_id']
        nome_produto = falta['nome_produto']
        qtd_falta = round(falta['estoque_minimo'] - falta['estoque_atual'], 2)
        media_destino = falta['media_diaria']

        todas_lojas = estoque_df[estoque_df['produto_id'] == produto_id]
        candidatos = todas_lojas[
            (todas_lojas['loja_id'] != loja_destino) &
            (todas_lojas['estoque_atual'] > todas_lojas['estoque_minimo']) &
            (todas_lojas['media_diaria'] <= media_destino + 0.5)
        ]

        if candidatos.empty:
            relatorio.append({
                "Loja com Falta": loja_destino,
                "Produto": nome_produto,
                "Quantidade Faltando": qtd_falta,
                "Loja Sugerida": "Nenhuma",
                "Quantidade Disponível": "",
                "Média de Vendas da Loja Sugerida": ""
            })
        else:
            for _, candidato in candidatos.iterrows():
                loja_origem = candidato['loja_id']
                qtd_sobrando = round(candidato['estoque_atual'] - candidato['estoque_minimo'], 2)
                media_origem = round(candidato['media_diaria'], 2)

                relatorio.append({
                    "Loja com Falta": loja_destino,
                    "Produto": nome_produto,
                    "Quantidade Faltando": qtd_falta,
                    "Loja Sugerida": loja_origem,
                    "Quantidade Disponível": qtd_sobrando,
                    "Média de Vendas da Loja Sugerida": media_origem
                })

df_relatorio = pd.DataFrame(relatorio)

excel_path = "relatorio_transferencias.xlsx"
df_relatorio.to_excel(excel_path, index=False)

wb = load_workbook(excel_path)
ws = wb.active

for cell in ws[1]:
    cell.font = Font(bold=True)

for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length + 2

wb.save(excel_path)

pdf = FPDF()
pdf.add_page()
pdf.set_font("Arial", "B", 14)
pdf.cell(0, 10, "Relatório de Estoque Crítico", ln=True, align="C")
pdf.ln(5)
pdf.set_font("Arial", size=11)

loja_atual = ""
for linha in relatorio:
    if linha["Loja com Falta"] != loja_atual:
        loja_atual = linha["Loja com Falta"]
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 10, f"Loja {loja_atual}", ln=True)
        pdf.set_font("Arial", size=11)

    produto = linha["Produto"]
    qtd_falta = linha["Quantidade Faltando"]
    loja_sug = linha["Loja Sugerida"]
    qtd_disp = linha["Quantidade Disponível"]
    media = linha["Média de Vendas da Loja Sugerida"]

    pdf.cell(0, 8, f"• {produto} - Faltando: {qtd_falta}", ln=True)

    if loja_sug != "Nenhuma":
        pdf.cell(0, 8, f"  → Sugerir da Loja {loja_sug}: {qtd_disp} unidades (média: {media})", ln=True)
    else:
        pdf.cell(0, 8, "  → Nenhuma loja sugerida", ln=True)

    pdf.ln(2)

pdf_path = "relatorio_transferencias.pdf"
pdf.output(pdf_path)

EMAIL_REMETENTE = os.getenv("EMAIL_REMETENTE")
EMAIL_DESTINATARIO = os.getenv("EMAIL_DESTINATARIO")
EMAIL_SENHA = os.getenv("EMAIL_SENHA")
SMTP_SERVIDOR = os.getenv("SMTP_SERVIDOR", "smtp.gmail.com")
SMTP_PORTA = int(os.getenv("SMTP_PORTA", 587))

msg = EmailMessage()
msg.set_content("Segue em anexo o relatório de estoque crítico com sugestões de transferência.")
msg["Subject"] = "Relatório de Estoque Crítico"
msg["From"] = EMAIL_REMETENTE
msg["To"] = EMAIL_DESTINATARIO

with open(excel_path, "rb") as f:
    msg.add_attachment(f.read(), maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=excel_path)

with open(pdf_path, "rb") as f:
    msg.add_attachment(f.read(), maintype="application", subtype="pdf", filename=pdf_path)

try:
    with smtplib.SMTP(SMTP_SERVIDOR, SMTP_PORTA) as server:
        server.starttls()
        server.login(EMAIL_REMETENTE, EMAIL_SENHA)
        server.send_message(msg)
    print("E-mail enviado com anexos com sucesso!")
except Exception as e:
    print("Erro ao enviar e-mail:", e)
